import sys, pandas as pd
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

OUT   = Path(r"D:\Geotab_EV_Parameters\charger_sizing_test\fixed_charger_milp_outputs")
EXCEL = OUT / "Fixed_Charger_MILP_Results.xlsx"
SITES = [("northgate","Northgate"),("fresno","Fresno"),("glendale","Glendale"),("san_diego","SanDiego")]

wb = load_workbook(EXCEL)
for _, lbl in SITES:
    sname = f"{lbl}_Worst10Sched"
    if sname in wb.sheetnames:
        del wb[sname]
wb.save(EXCEL)

with pd.ExcelWriter(EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace") as xl:
    for site, lbl in SITES:
        df = pd.read_csv(OUT / f"{site}_worst10_schedule.csv")
        df.to_excel(xl, sheet_name=f"{lbl}_Worst10Sched", index=False)
        max_soc = df["soc_end_pct"].replace("",None).dropna().astype(float).max()
        print(f"  {lbl}_Worst10Sched: {len(df)} rows  max_soc={max_soc:.1f}%")

print("Done.")

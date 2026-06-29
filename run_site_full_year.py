"""
run_site_full_year.py
──────────────────────
Full-year extended-dwell XOS analysis for a single Caltrans site.

Steps:
  1. Batch-extract per-day MILP event CSVs from the site Z2Z cache
     (skips days already on disk unless --overwrite is passed).
  2. Run extended-dwell XOS simulation on every extracted day.
  3. Save schedule .txt + Gantt .png per day.
  4. Save all-days summary CSV.
  5. Print percentile/cumulative coverage table.

Usage:
    python run_site_full_year.py --site fresno
    python run_site_full_year.py --site glendale
    python run_site_full_year.py --site san_diego
    python run_site_full_year.py --site san_diego --overwrite
    python run_site_full_year.py --site northgate --extract-only
    python run_site_full_year.py --site northgate --simulate-only

EV Matching priority chain (per vehicle):
  Tier 0   : device_name lookup in site vehicle master CSV (VIN-decoded authoritative)
  Tier 1   : EV_DIRECT_PATTERNS (vehicle is already a known EV)
  Tier 2   : EXTRA_ICE_OVERRIDES (hardcoded ICE -> EV)
  Tier 2.5 : Final Table make+model lookup (GVWR/segment-derived, covers all sites)
  Tier 3   : EV Equivalencies sheet exact/prefix match
  Tier 4   : difflib fuzzy match against EV Equivalencies sheet
"""
from __future__ import annotations

import sys, re, difflib, importlib, io, contextlib, argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Tuple, Set, Optional

import openpyxl
import pandas as pd
import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D

sys.stdout.reconfigure(encoding="utf-8")

# ── Site configuration ───────────────────────────────────────────────────────
BASE_DIR = Path(r"D:\Geotab_EV_Parameters\charger_sizing_test")
EV_CATEGORIES_XLSX = Path(r"D:\Geotab_EV_Parameters\final_categories.xlsx")
CHARGE_RATE_XLSX   = BASE_DIR / "ev_equivalent_max_charge_power_mapping_filled.xlsx"

VEHICLE_MASTER_DIR = Path(r"D:\Geotab_EV_Parameters")

SITE_CONFIG = {
    "northgate": {"cache": "_northgate_z2z_cache.csv",
                  "master": "northgate_vehicle_master.csv"},
    "fresno":    {"cache": "_fresno_z2z_cache.csv"},
    "glendale":  {"cache": "_glendale_z2z_cache.csv"},
    "san_diego": {"cache": "_san_diego_z2z_cache.csv"},
}

# ── Extraction parameters ────────────────────────────────────────────────────
MIN_DWELL_H    = 1.0
MIN_ENERGY_KWH = 0.10
TARGET_SOC     = 100.0
SOC_FALLBACK   = 50.0
ETA            = 0.90

EV_SPEC_OVERRIDES: dict[str, tuple[float, float]] = {
    "Ford F-150 Lightning":             (131.0,  0.4814),
    "Freightliner eCascadia":           (438.0,  2.10),
    "Freightliner eM2":                 (315.0,  1.164),
    "Tesla Model 3":                    (82.0,   0.259),
    "Rivian R1T":                       (135.0,  0.427),
    "Rivian R1S":                       (135.0,  0.427),
    "GMC Hummer EV":                    (212.7,  0.640),
    "BYD 6F Cab-Forward Truck":         (183.0,  1.540),
    "Chevrolet Silverado EV WT":        (200.0,  0.350),
    "Chevrolet Bolt EV":                (65.0,   0.281),
    "Kia EV6":                          (77.5,   0.288),
    "Blue Arc EV":                      (158.0,  1.000),
    "Ram ProMaster EV (cargo)":         (110.0,  0.671),
    "Ford E-Transit":                   (89.0,   0.560),
    "Volkswagen ID.4":                  (77.0,   0.347),
    "Volkswagen ID. Buzz":              (91.0,   0.406),
    "Volkswagen ID. Buzz (passenger)":  (91.0,   0.406),
    "Volvo VNR 4X2 Electric":           (375.0,  1.630),
    "Global Electric Street Sweeper (M4E)": (210.0, 4.421),
}

EV_DIRECT_PATTERNS: list[tuple[str, str]] = [
    ("tesla model 3",          "Tesla Model 3"),
    ("silverado ev",           "Chevrolet Silverado EV WT"),
    ("f-150 lightning",        "Ford F-150 Lightning"),
    ("ecascadia",              "Freightliner eCascadia"),
    ("em2",                    "Freightliner eM2"),
    ("rivian r1t",             "Rivian R1T"),
    ("rivian r1s",             "Rivian R1S"),
    ("hummer ev",              "GMC Hummer EV"),
    ("bolt ev",                "Chevrolet Bolt EV"),
    ("kia ev6",                "Kia EV6"),
    ("promaster ev",           "Ram ProMaster EV (cargo)"),
    ("e-transit",              "Ford E-Transit"),
    ("volkswagen id.4",        "Volkswagen ID.4"),
    ("volkswagen id. buzz",    "Volkswagen ID. Buzz"),
    ("id.4",                   "Volkswagen ID.4"),
    ("id. buzz",               "Volkswagen ID. Buzz"),
    ("blue arc",               "Blue Arc EV"),
    ("volvo vnr",              "Volvo VNR 4X2 Electric"),
    ("global electric sweeper","Global Electric Street Sweeper (M4E)"),
]

EXTRA_ICE_OVERRIDES: dict[str, str] = {
    # Heavy trucks
    "international hv":        "Freightliner eCascadia",
    "international hx":        "Freightliner eCascadia",
    "international workstar":  "Freightliner eCascadia",
    "international paystar":   "Freightliner eCascadia",
    "international 7600":      "Freightliner eCascadia",
    "international 5600":      "Freightliner eCascadia",
    "international 9400":      "Freightliner eCascadia",
    "international f-2674":    "Freightliner eCascadia",
    "western star":            "Volvo VNR 4X2 Electric",
    "freightliner 114 sd":     "Freightliner eCascadia",
    "freightliner fld":        "Freightliner eCascadia",
    "volvo vhd":               "Volvo VNR 4X2 Electric",
    "sterling l8500":          "Freightliner eCascadia",
    "sterling l9500":          "Freightliner eCascadia",
    "oshkosh h":               "Freightliner eCascadia",
    "autocar xpeditor":        "Freightliner eCascadia",
    "autocar wx":              "Freightliner eCascadia",
    # Medium trucks
    "freightliner m2":         "Freightliner eM2",
    "international durastar":  "Freightliner eM2",
    "international 5500":      "Freightliner eM2",
    "international 4700":      "Freightliner eM2",
    "international 4300":      "Freightliner eM2",
    "international 4400":      "Freightliner eM2",
    "gmc c series":            "Freightliner eM2",
    "gmc b7":                  "Freightliner eM2",
    "gmc c7":                  "Freightliner eM2",
    # Pickups / SUVs — not in Final Table, need explicit mapping
    "ford f-250":              "Ford F-150 Lightning",
    "ford f-350":              "GMC Hummer EV",
    "ford f-450":              "GMC Hummer EV",
    "ford f-550":              "GMC Hummer EV",
    "ford f-650":              "BYD 6F Cab-Forward Truck",
    "gmc sierra":              "Chevrolet Silverado EV WT",
    "gmc yukon":               "Rivian R1S",
    "chevrolet tahoe":         "Rivian R1S",
    "chevrolet suburban":      "Rivian R1S",
    "dodge durango":           "Rivian R1S",
    "dodge nitro":             "Rivian R1S",
    "nissan frontier":         "Rivian R1T",
    "ram 3500":                "GMC Hummer EV",
    "dodge ram pickup":        "GMC Hummer EV",
    "dodge ram chasis":        "GMC Hummer EV",
    "ram pickup heavy duty":   "GMC Hummer EV",
    "ram chasis cab":          "GMC Hummer EV",
    # Vans / cargo
    "ram promaster":           "Ram ProMaster EV (cargo)",
    "gmc savana":              "Ram ProMaster EV (cargo)",
    "ford windstar":           "Volkswagen ID. Buzz",
    # Compact / sedan (map to closest EV)
    "honda civic":             "Chevrolet Bolt EV",
    "pontiac vibe":            "Chevrolet Bolt EV",
    "toyota prius":            "Chevrolet Bolt EV",
    "ford expedition":         "Rivian R1S",
    "dodge charger":           "Tesla Model 3",
    # Explicitly exclude hydrogen / plug-in hybrid (return None via pattern list)
}

DC_FALLBACK: dict[str, float] = {"Global Electric Street Sweeper (M4E)": 60.0}
AC_FALLBACK: dict[str, float] = {"Global Electric Street Sweeper (M4E)": 0.0}


# ── EV matching ──────────────────────────────────────────────────────────────
def _load_ev_equivalencies(xlsx: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb["EV Equivalencies"]
    SKIP = {
        "ice example","iceexample","equivalent ev","category",
        "mpge city","mpge hwy","mpge comb","battery (kwh)","battery (kw)",
        "range (mi)","energy consumption at gvwr (kwh/mi)",
        "energy consumption (kwh/mi)","sweeping speed (mph)","sweeping time (h)",
    }
    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None
        if not isinstance(c1, str) or not isinstance(c2, str): continue
        k = c1.strip().lower()
        if k in SKIP or not k or not c2.strip(): continue
        out[k] = c2.strip()
    wb.close()
    return out


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def _norm_seg(s: str) -> str:
    """Normalize segment strings: collapse dashes, fix 'Trcuk' typo, lowercase."""
    return (re.sub(r"\s+", " ", s.strip().lower())
            .replace("–", "-").replace("—", "-")
            .replace("trcuk", "truck"))


def _load_vehicle_master(master_path: Path) -> dict[str, str]:
    """Load device_name → ev_equivalency from a vehicle master CSV."""
    if not master_path.exists():
        return {}
    df = pd.read_csv(str(master_path), low_memory=False)
    out: dict[str, str] = {}
    for _, r in df.iterrows():
        dname = str(r.get("device_name", "")).strip()
        ev    = str(r.get("ev_equivalency", "")).strip()
        if dname and ev and ev.lower() not in ("nan", "", "none"):
            out[dname] = ev
    return out


def _load_final_table_lookup(xlsx: Path) -> dict[tuple[str, str], str]:
    """
    Build (make_lower, model_lower) → ev_equivalency from the Final Table sheet.
    Uses EV Equivalencies sheet to translate Segment → EV equivalent.
    Handles combined Geotab model strings like "Silverado/Tahoe" by indexing
    each slash-separated part separately.
    """
    wb_ev = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    # Build segment → EV map from EV Equivalencies (col A = segment, col C = EV)
    seg_to_ev: dict[str, str] = {}
    ws_eq = wb_ev["EV Equivalencies"]
    for row in ws_eq.iter_rows(values_only=True):
        seg_raw = row[0] if len(row) > 0 else None
        ev_raw  = row[2] if len(row) > 2 else None
        if not isinstance(seg_raw, str) or not isinstance(ev_raw, str): continue
        seg = _norm_seg(seg_raw)
        ev  = ev_raw.strip()
        if seg and ev and seg not in ("nan", "ice example", "category"): seg_to_ev[seg] = ev
    wb_ev.close()

    # Load Final Table
    ft = pd.read_excel(str(xlsx), sheet_name="Final Table", header=0)
    ft["make_l"]  = ft["make"].fillna("").astype(str).str.strip().str.lower()
    ft["model_l"] = ft["model"].fillna("").astype(str).str.strip().str.lower()
    ft["seg_n"]   = ft["Segment"].fillna("").astype(str).apply(_norm_seg)
    ft["ev_eq"]   = ft["seg_n"].map(seg_to_ev)

    out: dict[tuple[str, str], str] = {}
    for _, r in ft.iterrows():
        ev = r["ev_eq"]
        if not isinstance(ev, str) or not ev.strip(): continue
        make_l  = r["make_l"]
        model_l = r["model_l"]
        if not make_l: continue
        # Index each slash-separated model part (handles "Silverado/Tahoe/...")
        for part in model_l.split("/"):
            part = part.strip()
            if not part: continue
            key = (make_l, part)
            if key not in out:
                out[key] = ev
    return out


def _match_ev(make: str, model: str, ice_to_ev: dict[str, str],
              vehicle_name: str | None = None,
              master_lookup: dict[str, str] | None = None,
              ft_lookup: dict[tuple[str, str], str] | None = None) -> str | None:
    # Tier 0: authoritative device_name lookup (VIN-decoded via merge_ev_equivalency)
    if vehicle_name and master_lookup:
        ev = master_lookup.get(str(vehicle_name).strip())
        if ev: return ev

    make, model = (make or "").strip(), (model or "").strip()
    if not make and not model: return None
    combo, model_n = _norm(f"{make} {model}"), _norm(model)

    # Tier 1: already an EV — detect by pattern
    for pat, ev in EV_DIRECT_PATTERNS:
        if pat in combo or pat in model_n: return ev

    # Tier 2: hardcoded ICE override table
    for pat, ev in EXTRA_ICE_OVERRIDES.items():
        if pat in combo or pat in model_n: return ev

    # Tier 2.5: Final Table make+model lookup (covers all Caltrans sites by GVWR/segment)
    if ft_lookup:
        make_l = _norm(make)
        for part in model_n.split("/"):
            part = part.strip()
            if not part: continue
            ev = ft_lookup.get((make_l, part))
            if ev: return ev

    # Tier 3: EV Equivalencies sheet — exact or prefix match on ICE model name
    if combo in ice_to_ev: return ice_to_ev[combo]
    if model_n in ice_to_ev: return ice_to_ev[model_n]
    for key, ev in ice_to_ev.items():
        if combo.startswith(key) or key.startswith(combo): return ev
        if model_n.startswith(key) or key.startswith(model_n): return ev

    # Tier 4: fuzzy match against EV Equivalencies sheet keys
    close = difflib.get_close_matches(combo, list(ice_to_ev), n=1, cutoff=0.60)
    if not close:
        close = difflib.get_close_matches(model_n, list(ice_to_ev), n=1, cutoff=0.60)
    return ice_to_ev[close[0]] if close else None


# ── Per-day extraction ───────────────────────────────────────────────────────
def extract_day(day_df: pd.DataFrame, date_str: str, site_slug: str,
                charge_map: dict, ice_to_ev: dict,
                master_lookup: dict | None = None,
                ft_lookup: dict | None = None) -> pd.DataFrame | None:
    unique_v = (day_df[["vehicle_name","make","model","year"]]
                .drop_duplicates("vehicle_name")
                .fillna({"make":"","model":"","year":""}))
    ev_rows = [{"vehicle_name": r["vehicle_name"],
                "ev_equivalent_model": _match_ev(
                    str(r["make"]), str(r["model"]), ice_to_ev,
                    vehicle_name=r["vehicle_name"],
                    master_lookup=master_lookup,
                    ft_lookup=ft_lookup)}
               for _, r in unique_v.iterrows()]
    ev_df = pd.DataFrame(ev_rows)
    ev_df = ev_df[ev_df["ev_equivalent_model"].notna()]
    df = day_df.merge(ev_df[["vehicle_name","ev_equivalent_model"]], on="vehicle_name", how="inner")
    if df.empty: return None

    df = df.copy()
    df["dwell_hours_actual"] = df["to_dwell_minutes"].fillna(0) / 60.0
    df["dwell_hours"] = df["dwell_hours_actual"].clip(lower=MIN_DWELL_H)
    short = df["dwell_hours_actual"] < MIN_DWELL_H
    df.loc[short, "to_exit_time"] = (df.loc[short, "to_entry_time"]
                                     + pd.to_timedelta(MIN_DWELL_H, unit="h"))

    def get_spec(ev, idx):
        s = EV_SPEC_OVERRIDES.get(ev); return s[idx] if s else float("nan")
    def get_ac(ev):
        cr = charge_map.get(ev); return float(cr["max_ac_charge_kw"]) if cr else AC_FALLBACK.get(ev, 0.0)
    def get_dc(ev):
        cr = charge_map.get(ev); return float(cr["max_dc_charge_kw"]) if cr else DC_FALLBACK.get(ev, 50.0)

    df["battery_capacity_kwh"]    = df["ev_equivalent_model"].map(lambda e: get_spec(e, 0))
    df["efficiency_kwh_per_mile"] = df["ev_equivalent_model"].map(lambda e: get_spec(e, 1))
    df["max_ac_charge_kw"]        = df["ev_equivalent_model"].map(get_ac)
    df["max_dc_charge_kw"]        = df["ev_equivalent_model"].map(get_dc)
    df = df[df["battery_capacity_kwh"].notna()].copy()
    if df.empty: return None

    dist = df["trip_first_distance_miles_between"].fillna(0).clip(lower=0)
    energy_used = dist * df["efficiency_kwh_per_mile"]
    arrival_soc = (SOC_FALLBACK - (energy_used / df["battery_capacity_kwh"] * 100.0)
                   ).clip(0.0, 100.0)
    df["assumed_initial_soc_percent"]  = arrival_soc.round(2)
    df["target_soc_percent"]           = TARGET_SOC
    df["energy_needed_kwh_for_visit"]  = (
        (TARGET_SOC - arrival_soc) / 100.0 * df["battery_capacity_kwh"]
    ).clip(lower=0.0).round(3)

    max_del = ETA * df["max_dc_charge_kw"].clip(upper=350) * df["dwell_hours"]
    df["individually_feasible"] = (df["energy_needed_kwh_for_visit"] <= max_del + MIN_ENERGY_KWH)

    svc = df[df["individually_feasible"]
             & (df["energy_needed_kwh_for_visit"] >= MIN_ENERGY_KWH)
             & df["energy_needed_kwh_for_visit"].notna()].copy().reset_index(drop=True)
    if svc.empty: return None

    svc = svc.sort_values("to_entry_time").reset_index(drop=True)
    date_tag = date_str.replace("-", "")
    svc["charging_event_id"] = [f"z2z_{date_tag}_v{i+1:02d}" for i in range(len(svc))]

    t_full_h = (svc["energy_needed_kwh_for_visit"] /
                (ETA * svc["max_dc_charge_kw"].clip(upper=350))).clip(lower=0)
    derived = svc[["dwell_hours_actual"]].assign(t=t_full_h).max(axis=1)
    miss = svc["to_exit_time"].isna()
    if miss.any():
        svc.loc[miss, "to_exit_time"] = (svc.loc[miss, "to_entry_time"]
                                          + pd.to_timedelta(derived[miss], unit="h"))
        svc.loc[miss, "dwell_hours"]        = derived[miss]
        svc.loc[miss, "dwell_hours_actual"] = derived[miss]

    svc["arrival_time"]   = svc["to_entry_time"].dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    svc["departure_time"] = svc["to_exit_time"].dt.strftime("%Y-%m-%dT%H:%M:%SZ")

    return svc[["charging_event_id","vehicle_name","arrival_time","departure_time",
                "dwell_hours","dwell_hours_actual","energy_needed_kwh_for_visit",
                "max_ac_charge_kw","max_dc_charge_kw","ev_equivalent_model",
                "individually_feasible","battery_capacity_kwh",
                "assumed_initial_soc_percent","target_soc_percent"]
               ].rename(columns={"vehicle_name": "vehicle_id"})


# ── Simulation helpers ───────────────────────────────────────────────────────
def _run_silent(events_df, p_eff, xos, MAX_U):
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        k, result = xos.find_min_xos_units(events_df, p_eff, max_units=MAX_U)
    return k, result


def compute_extensions(events_df, P_PORT, ETA_D):
    rows = []
    for _, row in events_df.iterrows():
        v      = row["charging_event_id"]
        arr    = row["arrival_time"]
        dep    = row["departure_time"]
        e_need = float(row["energy_needed_kwh_for_visit"])
        dwell_h  = (dep - arr).total_seconds() / 3600.0
        req_h    = e_need / (P_PORT * ETA_D)
        extra_h  = max(0.0, req_h - dwell_h)
        extended = extra_h > 1e-6
        ext_dep  = dep + pd.Timedelta(hours=extra_h) if extended else dep
        rows.append({"charging_event_id": v, "arrival_time": arr,
                     "original_departure_time": dep, "extended_departure_time": ext_dep,
                     "current_dwell_hours": round(dwell_h, 4),
                     "energy_needed_kwh": round(e_need, 3),
                     "required_dwell_hours_for_full_charge": round(req_h, 4),
                     "extra_dwell_hours_needed": round(extra_h, 4),
                     "was_dwell_extended": extended})
    ext_meta = pd.DataFrame(rows).set_index("charging_event_id")
    ext_events = events_df.copy()
    for idx, row in ext_events.iterrows():
        v = row["charging_event_id"]
        if ext_meta.loc[v, "was_dwell_extended"]:
            ext_events.at[idx, "departure_time"] = ext_meta.loc[v, "extended_departure_time"]
    return ext_events, ext_meta


def _dispatch_intervals(dispatch_log):
    ivs: dict = {}
    for entry in dispatch_log:
        v, t, k = entry["event_id"], pd.Timestamp(entry["time_utc"]), entry["unit"]
        if v not in ivs: ivs[v] = [t, t, {k}]
        else:
            if t < ivs[v][0]: ivs[v][0] = t
            if t > ivs[v][1]: ivs[v][1] = t
            ivs[v][2].add(k)
    DT = pd.Timedelta(minutes=15)
    return {v: (d[0], d[1] + DT, d[2]) for v, d in ivs.items()}


def write_gantt_png(events_df_orig, ext_meta, n_ext, res_ext,
                    out_path, date_str, site_label, xos):
    dispatch_ivs = _dispatch_intervals(res_ext["dispatch_log"])
    t_ref = events_df_orig["arrival_time"].min().floor("D")
    def hrs(ts): return (ts - t_ref).total_seconds() / 3600.0

    vehicles = sorted(events_df_orig["charging_event_id"].tolist(),
                      key=lambda v: ext_meta.loc[v, "arrival_time"])
    n_veh = len(vehicles)
    fig, ax = plt.subplots(figsize=(17, max(9, n_veh * 0.48 + 2.5)))
    BAR_H, CHARGE_H = 0.62, 0.36
    y_labels = []

    for i, v in enumerate(vehicles):
        m = ext_meta.loc[v]
        arr, odep, edep = m["arrival_time"], m["original_departure_time"], m["extended_departure_time"]
        y = i
        ax.barh(y, hrs(odep)-hrs(arr), left=hrs(arr), height=BAR_H,
                color="steelblue", alpha=0.30, edgecolor="steelblue", linewidth=0.6, zorder=2)
        if m["was_dwell_extended"] and m["extra_dwell_hours_needed"] > 1e-4:
            xtra_h = m["extra_dwell_hours_needed"]
            ax.barh(y, hrs(edep)-hrs(odep), left=hrs(odep), height=BAR_H,
                    color="darkorange", alpha=0.58, edgecolor="chocolate", linewidth=0.6, zorder=2)
            ax.text(hrs(odep)+(hrs(edep)-hrs(odep))/2, y, f"+{xtra_h:.2f}h",
                    ha="center", va="center", fontsize=6.8,
                    color="saddlebrown", fontweight="bold", zorder=5)
        if v in dispatch_ivs:
            t0, t1, _ = dispatch_ivs[v]
            t0p, t1p = max(t0, arr), min(t1, edep)
            if t1p > t0p:
                ax.barh(y, hrs(t1p)-hrs(t0p), left=hrs(t0p), height=CHARGE_H,
                        color="forestgreen", alpha=0.80, edgecolor="darkgreen",
                        linewidth=0.4, zorder=3)
        ax.plot([hrs(odep), hrs(odep)], [y-BAR_H/2, y+BAR_H/2],
                color="midnightblue", linewidth=0.9, linestyle="--", alpha=0.55, zorder=4)
        srv = "v" if v in dispatch_ivs else "x"
        v_short = str(v).rsplit("_", 1)[-1]
        y_labels.append(f"{srv} {v_short}  {m['energy_needed_kwh']:.0f} kWh")

    ax.set_yticks(range(n_veh))
    ax.set_yticklabels(y_labels, fontsize=8.2)
    ax.invert_yaxis()
    x_min = hrs(events_df_orig["arrival_time"].min()) - 0.4
    x_max = hrs(ext_meta["extended_departure_time"].max()) + 0.4
    ax.set_xlim(x_min, x_max)
    ticks = list(range(int(x_min), int(x_max)+2))
    ax.set_xticks(ticks)
    ax.set_xticklabels([f"{h%24:02d}:00" for h in ticks], fontsize=9)
    ax.xaxis.grid(True, linestyle=":", alpha=0.45, color="gray", zorder=0)
    ax.set_axisbelow(True)
    ax.tick_params(axis="y", length=0)
    ax.set_xlabel("Time of Day (UTC)", fontsize=10)
    ax.set_title(
        f"{site_label} {date_str} — XOS Extended Dwell Charging Schedule\n"
        f"{n_ext} XOS units  |  "
        f"{res_ext['n_served']}/{res_ext['n_total']} vehicles served (100% after extension)  |  "
        f"{res_ext['total_energy_delivered_kwh']:.0f} / {res_ext['total_energy_required_kwh']:.0f} kWh",
        fontsize=11, pad=10)
    ax.legend(handles=[
        mpatches.Patch(facecolor="steelblue", alpha=0.45, edgecolor="steelblue",
                       label="Original dwell window"),
        mpatches.Patch(facecolor="darkorange", alpha=0.7, edgecolor="chocolate",
                       label="Extended dwell"),
        mpatches.Patch(facecolor="forestgreen", alpha=0.85, edgecolor="darkgreen",
                       label="XOS charging period"),
        Line2D([0],[0], color="midnightblue", linewidth=1.2, linestyle="--",
               alpha=0.7, label="Original departure"),
    ], loc="lower right", fontsize=8.5, framealpha=0.92, edgecolor="gray")
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"      png: {out_path.name}")


def write_schedule_txt(events_df_orig, ext_meta, n_orig, res_orig,
                       n_ext, res_ext, out_path, date_str, site_label, xos, MAX_U):
    log, soc_hist = res_ext["dispatch_log"], res_ext["soc_history"]
    step_served: dict = defaultdict(lambda: defaultdict(list))
    step_kw: dict = defaultdict(float)
    for e in log:
        step_served[e["step_idx"]][e["unit"]].append(e["event_id"])
        step_kw[e["step_idx"]] += e["power_kw"]
    ext_rows = ext_meta[ext_meta["was_dwell_extended"]].sort_values("arrival_time")
    n_ext_veh = int(ext_meta["was_dwell_extended"].sum())
    avg_xtra  = ext_rows["extra_dwell_hours_needed"].mean() if n_ext_veh > 0 else 0.0
    max_xtra  = ext_meta["extra_dwell_hours_needed"].max()
    W, N_PORTS = 112, xos.N_PORTS
    L = [
        "=" * W,
        f"XOS HUB MC02 — EXTENDED DWELL CHARGING SCHEDULE ({site_label.upper()})",
        f"{'Date':<22}: {date_str}",
        f"{'Generated':<22}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "--- Baseline ---",
        f"  XOS units: {n_orig}  served: {res_orig['n_served']}/{res_orig['n_total']}"
        f"  ({100*res_orig['n_served']/max(res_orig['n_total'],1):.1f}%)",
        "",
        "--- Extended-Dwell ---",
        f"  XOS units: {n_ext}  served: {res_ext['n_served']}/{res_ext['n_total']}"
        f"  ({100*res_ext['n_served']/max(res_ext['n_total'],1):.1f}%)",
        f"  Vehicles extended: {n_ext_veh}/{len(ext_meta)}  avg +{avg_xtra:.2f}h  max +{max_xtra:.2f}h",
        "",
        "DWELL EXTENSIONS:",
        f"  {'Event ID':<26} {'Curr h':>7} {'Req h':>7} {'Extra h':>8}"
        f"  {'Orig dep':>18}  {'Ext dep':>18}  {'E need':>10}",
        f"  {'-'*104}",
    ]
    for v, r in ext_rows.iterrows():
        L.append(f"  {str(v):<26} {r['current_dwell_hours']:>7.2f} "
                 f"{r['required_dwell_hours_for_full_charge']:>7.2f} "
                 f"{r['extra_dwell_hours_needed']:>8.2f}"
                 f"  {str(r['original_departure_time'])[:16]:>18}"
                 f"  {str(r['extended_departure_time'])[:16]:>18}"
                 f"  {r['energy_needed_kwh']:>8.1f} kWh")
    L += ["", "-"*W, "STEP-BY-STEP DISPATCH (UTC)", "-"*W,
          f"  {'Time (UTC)':>16}  {'Vehicles'::<52}  {'kW':>7}  {'Ports':>6}  "
          + "  ".join(f"U{k}-SoC" for k in range(n_ext)), f"  {'-'*W}"]
    for sh in soc_hist:
        ti, t_str = sh["step_idx"], sh["time_utc"][:16].replace("T"," ")
        assignments = [f"{str(ev).rsplit('_',1)[-1]}->U{k}"
                       for k in range(n_ext) for ev in step_served[ti][k]]
        ports = sum(len(step_served[ti][k]) for k in range(n_ext))
        socs  = "  ".join(f"{sh.get(f'soc_unit_{k}',0.0):.3f}" for k in range(n_ext))
        L.append(f"  {t_str:>16}  {', '.join(assignments) or '- (grid recharge)':<52}"
                 f"  {step_kw[ti]:>7.1f}  {ports:>3}/{n_ext*N_PORTS}  {socs}")
    L.append("=" * W)
    out_path.write_text("\n".join(L), encoding="utf-8")
    print(f"      txt: {out_path.name}")


def process_day(csv_path, site_slug, site_label, OUT_BASE, xos, P_PORT, ETA_D, MAX_U):
    date_tag = csv_path.stem.split("_events_")[-1]          # e.g. "fresno_2025_05_01"
    date_str = date_tag[len(site_slug)+1:].replace("_", "-") # e.g. "2025-05-01"
    ext_dir  = OUT_BASE / f"xos_extended_{date_tag}"
    ext_dir.mkdir(parents=True, exist_ok=True)

    try:
        events_df = xos.load_events(csv_path)
    except Exception as exc:
        print(f"  [SKIP] {date_str}: {exc}")
        return None
    if len(events_df) == 0:
        return None

    p_eff = xos.compute_p_eff(events_df)
    n_orig, res_orig = _run_silent(events_df, p_eff, xos, MAX_U)

    ext_events, ext_meta = compute_extensions(events_df, P_PORT, ETA_D)
    p_eff_ext = xos.compute_p_eff(ext_events)
    n_units_ext, res_ext = _run_silent(ext_events, p_eff_ext, xos, MAX_U)

    ext_only  = ext_meta[ext_meta["was_dwell_extended"]]
    n_ext_veh = len(ext_only)
    avg_xtra  = float(ext_only["extra_dwell_hours_needed"].mean()) if n_ext_veh > 0 else 0.0
    max_xtra  = float(ext_meta["extra_dwell_hours_needed"].max())
    pct_b = 100 * res_orig["n_served"] / max(res_orig["n_total"], 1)
    pct_a = 100 * res_ext["n_served"]  / max(res_ext["n_total"],  1)
    print(f"  {date_str}  baseline={res_orig['n_served']}/{res_orig['n_total']} ({pct_b:.0f}%)"
          f"  extended={res_ext['n_served']}/{res_ext['n_total']} ({pct_a:.0f}%)"
          f"  units={n_units_ext}  avg+{avg_xtra:.2f}h")

    txt_path = ext_dir / f"xos_extended_schedule_{date_tag}.txt"
    png_path = ext_dir / f"xos_extended_schedule_{date_tag}.png"
    write_schedule_txt(events_df, ext_meta, n_orig, res_orig,
                       n_units_ext, res_ext, txt_path, date_str, site_label, xos, MAX_U)
    write_gantt_png(events_df, ext_meta, n_units_ext, res_ext,
                    png_path, date_str, site_label, xos)

    return {
        "date": date_str, "total_vehicles": res_orig["n_total"],
        "vehicles_extended": n_ext_veh,
        "avg_added_dwell_h": round(avg_xtra, 3),
        "max_added_dwell_h": round(max_xtra, 3),
        "served_before_extension": res_orig["n_served"],
        "served_after_extension":  res_ext["n_served"],
        "service_rate_before_pct": round(pct_b, 1),
        "service_rate_after_pct":  round(pct_a, 1),
        "min_xos_units_after_ext": n_units_ext,
        "all_served_after_ext":    res_ext["all_served"],
        "total_energy_required_kwh":  res_ext["total_energy_required_kwh"],
        "total_energy_delivered_kwh": res_ext["total_energy_delivered_kwh"],
        "peak_dispatch_kw":           res_ext["peak_dispatch_kw"],
    }


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--site", required=True, choices=list(SITE_CONFIG),
                        help="Site slug: northgate | fresno | glendale | san_diego")
    parser.add_argument("--overwrite",      action="store_true",
                        help="Overwrite existing per-day CSVs (extraction step)")
    parser.add_argument("--extract-only",   action="store_true",
                        help="Run extraction only; skip simulation")
    parser.add_argument("--simulate-only",  action="store_true",
                        help="Skip extraction; run simulation on existing per-day CSVs")
    args = parser.parse_args()

    site_slug  = args.site
    site_label = site_slug.replace("_", " ").title()
    cfg        = SITE_CONFIG[site_slug]
    cache_path = BASE_DIR / cfg["cache"]
    OUT_BASE   = BASE_DIR / "site_outputs" / site_slug
    OUT_BASE.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*72}")
    print(f"  {site_label.upper()} — FULL-YEAR XOS EXTENDED DWELL ANALYSIS")
    print(f"{'='*72}\n")

    # ── Step 1: Batch extraction ─────────────────────────────────────────────
    if not args.simulate_only:
        print("STEP 1: Batch Z2Z extraction")
        print(f"  Loading EV equivalencies and charge rates …")
        ice_to_ev  = _load_ev_equivalencies(EV_CATEGORIES_XLSX)
        cr_df = pd.read_excel(str(CHARGE_RATE_XLSX),
                              usecols=["ev_equivalent_model","max_ac_charge_kw","max_dc_charge_kw"])
        charge_map = (cr_df.drop_duplicates("ev_equivalent_model")
                      .set_index("ev_equivalent_model").to_dict("index"))

        # Tier 0: site-specific vehicle master (VIN-decoded by merge_ev_equivalency.py)
        master_file = cfg.get("master")
        master_path = VEHICLE_MASTER_DIR / master_file if master_file else None
        master_lookup = _load_vehicle_master(master_path) if master_path else {}
        if master_lookup:
            print(f"  Vehicle master loaded: {len(master_lookup)} device→EV mappings "
                  f"from {master_path.name}")
        else:
            print(f"  No vehicle master for {site_slug} (will use make/model matching only)")

        # Tier 2.5: Final Table make+model lookup (covers all Caltrans sites)
        print(f"  Building Final Table make+model lookup …")
        ft_lookup = _load_final_table_lookup(EV_CATEGORIES_XLSX)
        print(f"  Final Table lookup: {len(ft_lookup)} (make, model) entries")

        print(f"  Loading {cache_path.name} …")
        z2z = pd.read_csv(str(cache_path), low_memory=False)
        z2z = z2z[z2z["use_for_optimization_bool"].fillna(False).astype(bool)].copy()
        z2z["to_entry_time"] = pd.to_datetime(z2z["to_entry_time"], utc=True, errors="coerce")
        z2z["to_exit_time"]  = pd.to_datetime(z2z["to_exit_time"],  utc=True, errors="coerce")
        z2z = z2z.dropna(subset=["to_entry_time"])
        z2z["_date_pacific"] = (z2z["to_entry_time"]
                                .dt.tz_convert("America/Los_Angeles")
                                .dt.strftime("%Y-%m-%d"))

        all_dates = sorted(z2z["_date_pacific"].unique())
        print(f"  {len(z2z):,} events across {len(all_dates)} unique Pacific dates "
              f"({all_dates[0]} to {all_dates[-1]})")

        skipped = generated = failed = 0
        for date_str in all_dates:
            date_tag = date_str.replace("-", "_")
            out_path = BASE_DIR / f"z2z_milp_events_{site_slug}_{date_tag}.csv"
            if out_path.exists() and not args.overwrite:
                skipped += 1
                continue
            day_df = z2z[z2z["_date_pacific"] == date_str].copy()
            out_df = extract_day(day_df, date_str, site_slug, charge_map, ice_to_ev,
                                 master_lookup=master_lookup, ft_lookup=ft_lookup)
            if out_df is None or out_df.empty:
                failed += 1
                continue
            out_df.to_csv(str(out_path), index=False)
            generated += 1

        total_csvs = len(list(BASE_DIR.glob(f"z2z_milp_events_{site_slug}_*.csv")))
        print(f"  Generated: {generated}  Skipped: {skipped}  Empty: {failed}  "
              f"Total on disk: {total_csvs}\n")
    else:
        total_csvs = len(list(BASE_DIR.glob(f"z2z_milp_events_{site_slug}_*.csv")))
        print(f"STEP 1: Skipped (--simulate-only)  "
              f"[{total_csvs} existing CSVs on disk]\n")

    if args.extract_only:
        print("STEP 2: Skipped (--extract-only)")
        return

    # ── Step 2: Simulation ───────────────────────────────────────────────────
    print("STEP 2: XOS extended-dwell simulation")
    sys.path.insert(0, str(BASE_DIR))
    xos = importlib.import_module("xos_hub_soc_simulation")
    P_PORT = xos.P_PORT_KW
    ETA_D  = xos.ETA_D
    MAX_U  = xos.MAX_UNITS

    ALL_CSVS = sorted(BASE_DIR.glob(f"z2z_milp_events_{site_slug}_*.csv"))
    print(f"  Processing {len(ALL_CSVS)} days …\n")

    summary_rows = []
    for csv_path in ALL_CSVS:
        row = process_day(csv_path, site_slug, site_label, OUT_BASE,
                          xos, P_PORT, ETA_D, MAX_U)
        if row:
            summary_rows.append(row)

    # ── Save summary CSV ─────────────────────────────────────────────────────
    summary_df = pd.DataFrame(summary_rows)
    csv_out    = OUT_BASE / f"{site_slug}_extended_dwell_all_days_summary.csv"
    summary_df.to_csv(csv_out, index=False)
    print(f"\n  Summary CSV saved: {csv_out.name}")

    # ── Print results ────────────────────────────────────────────────────────
    tot_veh  = sum(r["total_vehicles"]             for r in summary_rows)
    tot_ext  = sum(r["vehicles_extended"]          for r in summary_rows)
    sbef     = sum(r["served_before_extension"]    for r in summary_rows)
    saft     = sum(r["served_after_extension"]     for r in summary_rows)
    e_req    = sum(r["total_energy_required_kwh"]  for r in summary_rows)
    e_del    = sum(r["total_energy_delivered_kwh"] for r in summary_rows)
    n100     = sum(1 for r in summary_rows if r["all_served_after_ext"])
    n_days   = len(summary_rows)
    units    = [r["min_xos_units_after_ext"] for r in summary_rows]

    print(f"\n{'='*72}")
    print(f"  {site_label.upper()} — FULL-YEAR SUMMARY")
    print(f"{'='*72}")
    print(f"  Operating days        : {n_days}")
    print(f"  Total vehicles        : {tot_veh:,}")
    print(f"  Vehicles extended     : {tot_ext:,} ({100*tot_ext/max(tot_veh,1):.1f}%)")
    print(f"  Avg added dwell       : {np.mean([r['avg_added_dwell_h'] for r in summary_rows if r['vehicles_extended']>0]):.2f} h")
    print(f"  Max added dwell       : {max(r['max_added_dwell_h'] for r in summary_rows):.2f} h")
    print(f"  Baseline service rate : {100*sbef/max(tot_veh,1):.1f}%")
    print(f"  Extended service rate : {100*saft/max(tot_veh,1):.1f}%  ({n100}/{n_days} days at 100%)")
    print(f"  Energy req / del      : {e_req:,.0f} / {e_del:,.0f} kWh")

    print(f"\n  Percentile fleet size (extended dwell):")
    print(f"  {'Percentile':<12} Units")
    print(f"  {'-'*20}")
    for p in [50, 70, 80, 90, 95, 99, 100]:
        print(f"  P{p:<11} {int(np.percentile(units, p))}")

    print(f"\n  Cumulative coverage:")
    print(f"  {'N units':<10} {'Days covered':<16} Coverage")
    print(f"  {'-'*36}")
    for n in sorted(set(units)):
        covered = sum(1 for u in units if u <= n)
        print(f"  {n:<10} {covered:>3} / {n_days:<11} {100*covered/n_days:.1f}%")

    print(f"\n  Done: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*72}\n")


if __name__ == "__main__":
    main()

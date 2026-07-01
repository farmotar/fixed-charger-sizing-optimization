"""
rebuild_worst10_schedules.py
Re-solves the 10 worst-cost days per site (from saved CSV rankings) and
regenerates per-vehicle schedule sheets with soc_end capped at 100%.
"""
from __future__ import annotations
import sys, importlib, warnings
from pathlib import Path
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, r"D:\Geotab_EV_Parameters\charger_sizing_test")

from charger_costs_caltrans import build_charger_specs_caltrans
milp = importlib.import_module("exact_northgate_charger_sizing_milp")

milp.OUTPUT_DIR         = Path(r"D:\Geotab_EV_Parameters\charger_sizing_test\fixed_charger_milp_outputs\_scratch")
milp.GUROBI_TIME_LIMIT  = 60
milp.GUROBI_MIP_GAP     = 0.05
milp.GUROBI_OUTPUT_FLAG = 0
milp.GUROBI_THREADS     = 0
milp.GUROBI_MIP_FOCUS   = 1
milp.DT_MINUTES         = 15
milp.DT_HOURS           = 15 / 60.0
milp.LAMBDA_SMOOTH      = 0.0

CHARGER_SPECS  = build_charger_specs_caltrans()
DAYS_PER_MONTH = 30.42
DAILY_CAPEX    = {
    ct: ((s["purchase_cost"] + s["install_cost"]) / (s["life_years"] * 12)
         + s["annual_maint"] / 12) / DAYS_PER_MONTH
    for ct, s in CHARGER_SPECS.items()
}

BASE  = Path(r"D:\Geotab_EV_Parameters\charger_sizing_test")
OUT   = BASE / "fixed_charger_milp_outputs"
EXCEL = OUT / "Fixed_Charger_MILP_Results.xlsx"

SITES = [
    ("northgate", "Northgate"),
    ("fresno",    "Fresno"),
    ("glendale",  "Glendale"),
    ("san_diego", "SanDiego"),
]


def local_time(ts, tz="America/Los_Angeles") -> str:
    if ts is None:
        return ""
    try:
        return ts.tz_convert(tz).strftime("%H:%M")
    except Exception:
        return ts.strftime("%H:%M")


def rebuild_site(site: str, site_label: str) -> pd.DataFrame:
    df_all     = pd.read_csv(OUT / f"{site}_all_days_milp.csv")
    worst10    = df_all[df_all["is_worst10"] == True].sort_values("cost_rank")
    worst_dates = worst10["date"].tolist()

    print(f"\n{'='*60}")
    print(f"  {site_label}  —  re-solving {len(worst_dates)} worst days")
    print(f"{'='*60}")

    all_rows = []

    for date in worst_dates:
        date_stem  = date.replace("-", "_")
        candidates = sorted(BASE.glob(f"z2z_milp_events_{site}_{date_stem}.csv"))
        if not candidates:
            print(f"  {date}: CSV not found, skipping")
            continue

        milp.RATE_SITE = site

        raw_df = pd.read_csv(candidates[0])
        ev_df  = milp.clean_events_df(raw_df)
        if ev_df is None or len(ev_df) == 0:
            print(f"  {date}: no valid events")
            continue

        tg = milp.build_time_grid(ev_df, dt_hours=milp.DT_HOURS)
        P  = milp.compute_effective_power(ev_df, CHARGER_SPECS)
        fk, E, _, _, avt = milp.build_feasible_keys(
            ev_df, tg, CHARGER_SPECS, P, dt_hours=milp.DT_HOURS
        )
        cd, cp = milp.site_capacity_charge_rates(site)

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            sol = milp.solve_with_gurobi(
                events_df         = ev_df,
                time_grid         = tg,
                charger_specs     = CHARGER_SPECS,
                daily_capex       = DAILY_CAPEX,
                P_eff             = P,
                feasible_keys     = fk,
                E                 = E,
                available_times   = avt,
                dt_hours          = milp.DT_HOURS,
                eta               = milp.ETA,
                c_demand_global   = cd,
                c_demand_peak_win = cp,
                lambda_smooth       = 0.0,
                lambda_energy_error = milp.LAMBDA_ENERGY_ERROR,
            )

        if sol.get("status") in ("infeasible", "no_solution", None):
            print(f"  {date}: solver returned {sol.get('status')} — skip")
            continue

        # ── Extract delivery from x_vals ──────────────────────────────────
        x_vals       = sol.get("x_vals", {})
        delivered    = {v: 0.0  for v in E}
        charge_start = {v: None for v in E}
        charge_end   = {v: None for v in E}
        charger_type = {v: ""   for v in E}

        for (v, tidx, c), pwr in x_vals.items():
            delivered[v] += pwr * milp.DT_HOURS * milp.ETA
            t    = tg[tidx]
            t_end = t + pd.Timedelta(hours=milp.DT_HOURS)
            if charge_start[v] is None:
                charge_start[v] = t
            charge_end[v]   = t_end
            charger_type[v] = c

        ev_lk    = ev_df.set_index("charging_event_id")
        day_info = worst10[worst10["date"] == date].iloc[0]

        n_full = n_partial = n_unserved = 0
        for v in E:
            re_v   = ev_lk.loc[v] if v in ev_lk.index else {}
            needed = E[v]
            deliv  = delivered[v]

            if deliv >= needed - milp.ENERGY_TOL:
                status = "full";    n_full     += 1
            elif deliv > milp.ENERGY_TOL:
                status = "partial"; n_partial  += 1
            else:
                status = "unserved"; n_unserved += 1

            arr_utc = pd.Timestamp(re_v["arrival_time"])   if "arrival_time"   in re_v else None
            dep_utc = pd.Timestamp(re_v["departure_time"]) if "departure_time" in re_v else None
            dwell   = (dep_utc - arr_utc).total_seconds() / 3600 if (arr_utc and dep_utc) else 0.0

            cs  = charge_start[v]
            ce  = charge_end[v]
            dur = (ce - cs).total_seconds() / 3600 if (cs and ce) else 0.0

            soc_start_pct = float(re_v.get("assumed_initial_soc_percent", 0) or 0)
            bat_cap       = float(re_v.get("battery_capacity_kwh", 0) or 0)
            if bat_cap > 0:
                # Cap at 100% — 15-min slots can deliver slightly more than battery room
                soc_end_pct = min(100.0, soc_start_pct + 100.0 * deliv / bat_cap)
            else:
                soc_end_pct = None

            all_rows.append({
                "date":                 date,
                "worst_day_rank":       int(day_info["cost_rank"]),
                "total_op_cost":        round(float(day_info["total_op_cost"]), 2),
                "config_label":         day_info["config_label"],
                "charging_event_id":    v,
                "vehicle_id":           str(re_v.get("vehicle_id", "")),
                "ev_model":             str(re_v.get("ev_equivalent_model", "")),
                "arrival_local":        local_time(arr_utc),
                "departure_local":      local_time(dep_utc),
                "dwell_h":              round(dwell, 2),
                "energy_needed_kwh":    round(needed, 1),
                "energy_delivered_kwh": round(deliv, 1),
                "energy_gap_kwh":       round(max(needed - deliv, 0), 1),
                "status":               status,
                "charger_type_used":    charger_type[v],
                "charge_start":         local_time(cs),
                "charge_end":           local_time(ce),
                "charge_duration_h":    round(dur, 2),
                "soc_start_pct":        round(soc_start_pct, 1),
                "soc_end_pct":          round(soc_end_pct, 1) if soc_end_pct is not None else "",
            })

        print(f"  {date}  rank#{int(day_info['cost_rank'])}  "
              f"full={n_full} partial={n_partial} unserved={n_unserved}  "
              f"svc={100*(n_full+n_partial)/max(len(E),1):.0f}%")

    df = pd.DataFrame(all_rows)
    if len(df):
        df = df.sort_values(["worst_day_rank", "arrival_local"]).reset_index(drop=True)

    # Validate: no soc_end > 100
    bad = df[df["soc_end_pct"].apply(lambda x: isinstance(x, float) and x > 100)]
    if len(bad):
        print(f"  [WARNING] {len(bad)} rows still have soc_end > 100 — investigate!")
    else:
        max_soc = df["soc_end_pct"].replace("", None).dropna().astype(float).max()
        print(f"  Max soc_end_pct = {max_soc:.1f}%  (all <= 100 OK)")

    csv_out = OUT / f"{site}_worst10_schedule.csv"
    df.to_csv(csv_out, index=False)
    print(f"  Saved: {csv_out.name}  ({len(df)} vehicle-day rows)")
    return df


# ── Run all sites ─────────────────────────────────────────────────────────────
site_dfs = {}
for site, site_label in SITES:
    site_dfs[site_label] = rebuild_site(site, site_label)

# ── Rewrite schedule sheets in Excel ─────────────────────────────────────────
print("\nUpdating Excel schedule sheets...")
from openpyxl import load_workbook

wb = load_workbook(EXCEL)
for site_label in site_dfs:
    sname = f"{site_label}_Worst10Sched"
    if sname in wb.sheetnames:
        del wb[sname]
wb.save(EXCEL)

with pd.ExcelWriter(EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace") as xl:
    for site_label, df in site_dfs.items():
        sheet = f"{site_label}_Worst10Sched"
        df.to_excel(xl, sheet_name=sheet, index=False)
        print(f"  Written: {sheet}")

print(f"\nDone. Updated: {EXCEL}")
